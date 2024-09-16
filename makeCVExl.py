import requests
import json
import time
import os
import openpyxl
import csv
import glob
import logging
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed

# ロガーの設定
logger = logging.getLogger()
logger.setLevel(logging.INFO)
file_handler = logging.FileHandler('result.log', mode='w')
file_handler.setLevel(logging.INFO)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# 定数
CACHE_PATH = "cache"
XLSX_FILENAME = "result.xlsx"
SHEET_NAME = XLSX_FILENAME.replace(".xlsx", "")

COLUMN_INDEXES = {
    'PluginID': 1,
    'Risk': 2,
    'Name': 3,
    'CVE': 4,
    'CWE': 5,
    'V20BS': 6,
    'V20Vector': 7,
    'V30BS': 8,
    'V30Vector': 9,
    'V31BS': 10,
    'V31Vector': 11,
    'V40BS': 12,
    'V40Vector': 13
}

def fill_row_with_color(sheet, row_number, fill):
    # 指定された行のすべてのセルに背景色を設定します。
    for cell in sheet[row_number]:
        cell.fill = fill

def getCVEInfo(CVEID):
    # CVE ID に基づいて CVE 情報を取得し、キャッシュディレクトリに保存します。
    filename = os.path.join(CACHE_PATH, f"{CVEID}.json")
    if not os.path.isdir(CACHE_PATH):
        logger.info(f"Creating the folder named {CACHE_PATH}.")
        os.makedirs(CACHE_PATH)
    
    if not os.path.isfile(filename):
        logger.info(f"Fetching CVE data for {CVEID}.")
        attempt = 0
        max_attempts = 3
        while attempt < max_attempts:
            try:
                url = f"https://services.nvd.nist.gov/rest/json/cves/2.0?cveId={CVEID}"
                headers = {"Accept": "application/json"}
                res = requests.get(url, headers=headers)
                res.raise_for_status()
                with open(filename, mode="w", encoding="utf-8") as f:
                    json.dump(res.json(), f, indent=4)
                logger.info(f"Saved JSON file: {filename}")
                break
            except requests.RequestException as e:
                attempt += 1
                logger.error(f"API request failed: {e}. Attempt {attempt} of {max_attempts}")
                if attempt < max_attempts:
                    time.sleep(1)  # Exponential backoff could be applied here
                else:
                    logger.error(f"Failed to fetch CVE data after {max_attempts} attempts.")
            except IOError as e:
                logger.error(f"File I/O error: {e}")
                break
    else:
        logger.info(f"File {filename} already exists.")

def write_to_excel(basename, cvedict):
    # CVE データを Excel ファイルに書き込みます。
    file_path = f"{basename}_{XLSX_FILENAME}"
    if os.path.isfile(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[SHEET_NAME]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        # 列のヘッダーを設定
        for col_name, col_idx in COLUMN_INDEXES.items():
            ws.cell(1, col_idx).value = col_name

    # 書き込むデータを収集
    rows_to_write = []
    for row_idx, (cveid, row_data) in enumerate(cvedict.items(), start=2):
        row = [
            row_data.get('Plugin ID', ''),
            row_data.get('Risk', ''),
            row_data.get('Name', ''),
            cveid
        ]
        rows_to_write.append(row)
    
    # すべてのデータを一度に Excel に書き込み
    for row_idx, row_data in enumerate(rows_to_write, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row_idx, col_idx).value = value

    wb.save(file_path)

def update_excel_with_cve_info(basename):
    # Excel ファイルを更新し、CVE 情報を追加します。
    file_path = f"{basename}_{XLSX_FILENAME}"
    if not os.path.isfile(file_path):
        logger.error(f"Excel file {file_path} does not exist.")
        return
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb[SHEET_NAME]
    
    # Excel シートから CVE ID のリストを取得
    cve_ids = [row[COLUMN_INDEXES['CVE'] - 1].value for row in ws.iter_rows(min_row=2)]
    
    # CVE 情報を並行して取得
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_cve = {executor.submit(getCVEInfo, cve_id): cve_id for cve_id in cve_ids}
        for future in as_completed(future_to_cve):
            cveid = future_to_cve[future]
            try:
                future.result()
            except Exception as e:
                logger.error(f"Error fetching CVE data for {cveid}: {e}")
    
    # Excel シートの各行を更新
    for row in ws.iter_rows(min_row=2):
        cveid = row[COLUMN_INDEXES['CVE'] - 1].value
        filename = os.path.join(CACHE_PATH, f"{cveid}.json")
        
        if not os.path.isfile(filename):
            logger.warning(f"Cache file {filename} does not exist.")
            continue
        
        try:
            with open(filename) as jfile:
                json_dict = json.load(jfile)
                vulnerabilities = json_dict.get("vulnerabilities", [])
                if not vulnerabilities:
                    logger.warning(f"No vulnerabilities found in {filename}")
                    continue

                cvss_dicts = vulnerabilities[0].get("cve", {}).get("metrics", {})
                cve_info = vulnerabilities[0].get("cve", {})
                if "weaknesses" in cve_info:
                    ws.cell(row[0].row, COLUMN_INDEXES['CWE']).value = cve_info["weaknesses"][0]["description"][0]["value"]
                
                CVSSScores = {key: "N/A" for key in COLUMN_INDEXES.keys() if key.startswith('V')}
                for cvss_key, cvss_label in [('cvssMetricV2', 'V20'), ('cvssMetricV30', 'V30'), ('cvssMetricV31', 'V31'), ('cvssMetricV40', 'V40')] :
                    if cvss_key in cvss_dicts:
                        cvss_data = cvss_dicts[cvss_key][0]["cvssData"]
                        CVSSScores[f"{cvss_label}BS"] = str(cvss_data["baseScore"])
                        CVSSScores[f"{cvss_label}Vector"] = str(cvss_data["vectorString"]).replace(f"CVSS:{cvss_label}/", "")
                        ws.cell(row[0].row, COLUMN_INDEXES[f'{cvss_label}BS']).value = CVSSScores[f"{cvss_label}BS"]
                        ws.cell(row[0].row, COLUMN_INDEXES[f'{cvss_label}Vector']).value = CVSSScores[f"{cvss_label}Vector"]
                
                CVSSBaseScores = {k: v for k, v in CVSSScores.items() if 'BS' in k}
                if all(score == "N/A" for score in CVSSBaseScores.values()):
                    logger.error(f"No CVSS BaseScores found for {cveid}! Highlighting row.")
                    fill_row_with_color(ws, row[0].row, PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'))
                else:
                    logger.info(f"The CVSS BaseScores for {cveid} are: {', '.join(f'{key}: {value}' for key, value in CVSSBaseScores.items())}")
        except (json.JSONDecodeError, IOError) as e:
            logger.error(f"Error processing JSON file {filename}: {e}")

    try:
        wb.save(file_path)
    except Exception as e:
        logger.error(f"Error saving Excel file {file_path}: {e}")

def process_csv_file(file):
    # CSV ファイルを処理し、CVE データを Excel ファイルに書き込みます。
    basename = os.path.splitext(os.path.basename(file))[0]
    try:
        if not os.path.isfile(file):
            logger.error(f"CSV file {file} does not exist.")
            return

        with open(file, encoding='utf-8') as csvf:
            reader = csv.DictReader(csvf)
            cvedict = {row['CVE']: row for row in reader if row['CVE']}
        
        write_to_excel(basename, cvedict)
        update_excel_with_cve_info(basename)
    except Exception as e:
        logger.error(f"Error processing file {file}: {e}")

def main():
    # CSV フォルダ内のすべての CSV ファイルを処理します。
    logger.info("Started")
    csv_dir = "csv"
    if os.path.isdir(csv_dir) and os.listdir(csv_dir):
        files = glob.glob(os.path.join(csv_dir, "*.csv"))
        for file in files:
            logger.info(f"Processing CSV file: {os.path.basename(file)}")
            process_csv_file(file)
    elif not os.path.isdir(csv_dir):
        logger.error(f"The folder named '{csv_dir}' was not found. Creating the folder.")
        os.mkdir(csv_dir)
        logger.error("Please move the CSV file exported from Nessus to the folder.")
    else:
        logger.critical("No CSV files found in the 'csv' folder. Please move the CSV file exported from Nessus to the folder.")
    logger.info("Finished")

if __name__ == '__main__':
    main()
